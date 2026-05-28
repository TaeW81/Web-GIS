"""
엑셀 내보내기 모듈 (안정성 극대화 버전)
"""
import io
import re
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

def format_land_ledger_sheet(ws):
    """토지조서 시트에 스타일 적용 (범위 안전성 확보)"""
    if ws.max_row < 1: return
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 열 너비 설정 (A~X, 24개 컬럼 — W=토지보상비 추가, 비고는 X로 이동)
    ws.column_dimensions['A'].width = 8   # 일련번호
    ws.column_dimensions['B'].width = 22  # PNU
    ws.column_dimensions['C'].width = 30  # 소재지
    ws.column_dimensions['D'].width = 10  # 필지구분
    ws.column_dimensions['E'].width = 6   # 본번
    ws.column_dimensions['F'].width = 6   # 부번
    ws.column_dimensions['G'].width = 8   # 지목
    ws.column_dimensions['H'].width = 12  # 소유자
    ws.column_dimensions['I'].width = 8   # 소유자수
    ws.column_dimensions['J'].width = 12  # 공시지가
    ws.column_dimensions['K'].width = 12  # 대장면적
    ws.column_dimensions['L'].width = 12  # 편입면적
    ws.column_dimensions['M'].width = 10  # 편입구분
    ws.column_dimensions['N'].width = 25  # 용도지역
    ws.column_dimensions['O'].width = 12  # 진흥지역
    ws.column_dimensions['P'].width = 10  # 농지구분
    ws.column_dimensions['Q'].width = 10  # 산지구분
    ws.column_dimensions['R'].width = 15  # 이용상황
    # 보상조서 영역 (S~V)
    ws.column_dimensions['S'].width = 18  # 용도지역(보상)
    ws.column_dimensions['T'].width = 10  # 보상배율(용도)
    ws.column_dimensions['U'].width = 14  # 이용상황(보상)
    ws.column_dimensions['V'].width = 10  # 보상배율(이용)
    # 토지보상비 + 비고
    ws.column_dimensions['W'].width = 18  # 토지보상비(원)
    ws.column_dimensions['X'].width = 20  # 비고

    max_col = ws.max_column or 24
    # pandas to_excel 직후: 행 1=헤더, 행 2~last=데이터
    n_data = ws.max_row - 1  # 헤더 제외 데이터 행 개수

    # ★ 합계 행을 데이터 위(행 2)에 삽입 — 헤더 바로 아래에 표기
    total_row_idx = None
    if n_data >= 1:
        ws.insert_rows(2)              # 행 2에 빈 행 삽입 → 데이터는 행 3~로 이동
        total_row_idx = 2
        last_data_row = n_data + 2     # 데이터 마지막 행 (= 2 + n_data)

        # 합계 라벨 + SUM 수식 (K=대장면적, L=편입면적, W=토지보상비)
        ws.cell(row=2, column=1, value="합계")
        ws.cell(row=2, column=11,
                value=f'=SUM(K3:K{last_data_row})').number_format = '#,##0.00'  # K
        ws.cell(row=2, column=12,
                value=f'=SUM(L3:L{last_data_row})').number_format = '#,##0.00'  # L
        ws.cell(row=2, column=23,
                value=f'=SUM(W3:W{last_data_row})').number_format = '#,##0'      # W
    else:
        last_data_row = ws.max_row     # 데이터 없음

    # ★ W열(토지보상비) 수식 작성 (데이터 행 3~last_data_row)
    #   - 국공유지(U="국ㆍ공유지") → 0
    #   - 공시지가(J) 숫자 아님 또는 0 → 0
    #   - 그 외 → J × L × (T+V)/2  (T 또는 V가 빈셀이면 IFERROR로 0 처리)
    #   출처: 예타 총괄지침 — 보상배율 = (용도지역 배율 + 이용상황 배율) / 2
    data_start = 3 if total_row_idx else 2
    for r in range(data_start, last_data_row + 1):
        formula = (
            f'=IFERROR('
            f'IF(OR(U{r}="국ㆍ공유지", NOT(ISNUMBER(J{r})), J{r}=0), 0, '
            f'J{r}*L{r}*(T{r}+V{r})/2)'
            f', 0)'
        )
        ws.cell(row=r, column=23, value=formula)  # W = 23번째 컬럼
        ws.cell(row=r, column=23).number_format = '#,##0'  # 천단위 콤마

    # 공시지가(J)에도 천단위 콤마 (가독성)
    for r in range(data_start, last_data_row + 1):
        j_cell = ws.cell(row=r, column=10)
        if isinstance(j_cell.value, (int, float)):
            j_cell.number_format = '#,##0'

    # 합계 행 강조 서식 (옅은 노랑 + 굵게)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_font = Font(bold=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col), 1):
        for cell in row:
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            elif total_row_idx and row_idx == total_row_idx:
                cell.fill = total_fill; cell.font = total_font; cell.alignment = center_align
            else:
                if cell.column_letter in ['C', 'N', 'S']:  # 소재지, 용도지역, 용도지역(보상)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = center_align

def create_multi_sheet_excel(sheets_data):
    """IndexError를 원천 차단하는 가장 단순하고 확실한 방법"""
    buffer = io.BytesIO()
    
    # 데이터 유효성 검사 및 보정
    valid_data = {k: v for k, v in (sheets_data or {}).items() if v}
    if not valid_data:
        valid_data = {"결과없음": [{"상태": "분석된 데이터가 없습니다."}]}

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for i, (sheet_name, data_list) in enumerate(valid_data.items()):
            df = pd.DataFrame(data_list)
            
            # ⭐️ 시트명 금지 문자 처리 (/, \, ?, *, [, ], :) ⭐️
            clean_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
            
            df.to_excel(writer, index=False, sheet_name=clean_sheet_name)
            
            if "토지조서" in clean_sheet_name:
                format_land_ledger_sheet(writer.sheets[clean_sheet_name])

        # ⭐️ 핵심: 모든 작업이 끝난 후, 'Sheet'나 'Sheet1'이 남아있고 다른 시트가 있다면 삭제 ⭐️
        # 삭제 후에는 반드시 active 시트를 설정해야 함
        workbook = writer.book
        for default_sheet in ["Sheet", "Sheet1"]:
            if default_sheet in workbook.sheetnames and len(workbook.sheetnames) > 1:
                workbook.remove(workbook[default_sheet])
        
        # 마지막 안전장치: 첫 번째 시트를 활성 시트로 지정
        workbook.active = 0

    return buffer.getvalue()
