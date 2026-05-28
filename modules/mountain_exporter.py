"""
편입산지조서 생성 모듈

샘플: 연습용자료/편입산지조서(sample).xlsx
템플릿: assets/편입산지조서_template.xlsx

요구사항:
  - 산지 필지(지목="임" 또는 필지구분="산")만 추출
  - 각 필지당 "당초" 1행씩 작성 (변경/증감 행은 제거)
  - 상단 합계 영역도 "당초"만 유지 (변경/증감 합계 행 제거)
  - 합계 행의 SUMIFS는 단순 SUM으로 단순화 (B열 필터 제거)
"""
import os
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from copy import copy

# 데이터 영역에 강제 적용할 표준 서식 (unmerge 후 깨진 셀 복구용)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 빈 채우기 (파란색 등 잔존 색 제거용)
NO_FILL = PatternFill(fill_type=None)

# 지목 한글 명칭 → 1글자 약어 (편입산지조서 F열 간소화용)
JIMOK_SHORT_MAP = {
    "전": "전", "답": "답", "과수원": "과", "목장용지": "목", "임야": "임",
    "광천지": "광", "염전": "염", "대": "대", "공장용지": "장",
    "학교용지": "학", "주차장": "주", "창고용지": "창", "도로": "도",
    "철도용지": "철", "제방": "제", "구거": "구", "유지": "유",
    "수도용지": "수", "공원": "공", "체육용지": "체", "유원지": "원",
    "종교용지": "종", "사적지": "사", "묘지": "묘", "잡종지": "잡",
    # 1글자가 들어오는 경우(이미 짧음)도 안전하게 통과
    "과": "과", "목": "목", "임": "임", "장": "장", "학": "학",
    "주": "주", "창": "창", "도": "도", "철": "철", "제": "제",
    "구": "구", "유": "유", "수": "수", "공": "공", "체": "체",
    "원": "원", "종": "종", "사": "사", "묘": "묘", "잡": "잡",
    "광": "광", "염": "염",
}


def _short_jimok(jimok_raw):
    """지목 한글 명칭을 1글자 약어로 변환."""
    if not jimok_raw:
        return ""
    j = str(jimok_raw).strip()
    return JIMOK_SHORT_MAP.get(j, j[0] if j else j)


def _classify_owner(owner_type):
    """소유자 표기 → 국유지/공유지/사유지 분류."""
    if not owner_type or owner_type == "-":
        return "사유지"
    if "국유" in owner_type:
        return "국유지"
    if any(k in owner_type for k in ["군유", "도유", "시유", "공유", "시도유"]):
        return "공유지"
    return "사유지"  # 개인, 법인, 종중, 기타단체, 공공기관 등


def _split_mountain_zone(mountain_type, inc_area):
    """산지구분 텍스트 → (임업용/공익용/준보전) 면적 배분.
    현재 analyzer는 "산지" 또는 "-"만 반환하므로 대부분 임업용으로 분류됨.
    추후 산림청 산지구분 API 연동 시 세분화될 예정."""
    if not mountain_type or mountain_type in ("-", ""):
        return 0.0, 0.0, 0.0
    z = str(mountain_type)
    if "공익" in z:
        return 0.0, inc_area, 0.0
    if "준보전" in z:
        return 0.0, 0.0, inc_area
    # "보전", "임업", "산지" 등은 기본적으로 임업용 보전산지
    return inc_area, 0.0, 0.0


def _extract_admin_address(sojaeji_full):
    """소재지 풀텍스트에서 시/군/구/읍/면/동/리만 추출.
    도 단위(경기도/경상북도 등)는 제외 — 사용자 요청 ("광명시 가학동" 형식)."""
    if not sojaeji_full:
        return ""
    parts = str(sojaeji_full).split()
    pieces = [
        p for p in parts
        if any(p.endswith(suf) for suf in ("시", "군", "구", "읍", "면", "동", "리"))
        # 단, "광역시"/"특별시" 같이 "시"로 끝나는 광역시는 자연스럽게 포함됨
    ]
    return " ".join(pieces) if pieces else sojaeji_full


def _format_jibun(parcel):
    """본번/부번 → '본번-부번' 형식, 산번지면 앞에 '산' 추가."""
    bonbun = str(parcel.get("본번", "0") or "0")
    bubun = str(parcel.get("부번", "0") or "0")
    jibun = bonbun if bubun in ("0", "", None) else f"{bonbun}-{bubun}"
    if parcel.get("필지구분") == "산":
        jibun = f"산 {jibun}"
    return jibun


def _fill_data_row(ws, row, idx, parcel):
    """편입산지조서 데이터 1행 채우기 (한 필지 당 1행).

    합계 셀(J/N/O/T/X/Y)은 옆 데이터 셀 합계 수식으로 작성:
      J = K + L + M             (원형존치 소유별 합)
      N = O + R                 (원형존치 산지구분별 합)
      O = P + Q                 (원형존치 보전산지 소계)
      T = U + V + W             (산지전용협의 소유별 합)
      X = Y + AB                (산지전용협의 산지구분별 합)
      Y = Z + AA                (산지전용협의 보전산지 소계)
    """
    sojaeji = _extract_admin_address(parcel.get("소재지", ""))
    jibun = _format_jibun(parcel)
    jimok = _short_jimok(parcel.get("지목", "임"))  # "임야" → "임" 등 1글자 약어
    owner = parcel.get("소유자", "")
    cad_area = float(parcel.get("대장면적(㎡)", 0.0) or 0.0)
    inc_area = float(parcel.get("편입면적(㎡)", 0.0) or 0.0)

    # 소유자 분류 (국유/공유/사유)
    own_class = _classify_owner(owner)
    val_guk = inc_area if own_class == "국유지" else 0.0
    val_gong = inc_area if own_class == "공유지" else 0.0
    val_sa = inc_area if own_class == "사유지" else 0.0

    # 산지구분 분류
    val_imup, val_gongik, val_junbo = _split_mountain_zone(parcel.get("산지구분", ""), inc_area)

    # 기본 정보 컬럼
    ws.cell(row=row, column=1, value=idx)              # A: 연번
    # B: 구분 — 단일 행이므로 비움
    ws.cell(row=row, column=3, value=sojaeji)          # C: 소재지
    ws.cell(row=row, column=5, value=jibun)            # E: 지번
    ws.cell(row=row, column=6, value=jimok)            # F: 지목
    ws.cell(row=row, column=7, value=cad_area)         # G: 공부면적
    ws.cell(row=row, column=8, value=inc_area)         # H: 편입면적 (① 구역지정)

    # ② 원형존치 (현재 미사용 — 모두 0)
    # I: ② 원형존치 면적 = J (소유별 합) (수식)
    ws.cell(row=row, column=9, value=f"=J{row}")
    # J: 원형존치 소유별 합계 = K + L + M (수식)
    ws.cell(row=row, column=10, value=f"=K{row}+L{row}+M{row}")
    ws.cell(row=row, column=11, value=0)               # K: 원형존치 국유지
    ws.cell(row=row, column=12, value=0)               # L: 원형존치 공유지
    ws.cell(row=row, column=13, value=0)               # M: 원형존치 사유지
    # N: 원형존치 산지구분별 합계 = O + R (수식)
    ws.cell(row=row, column=14, value=f"=O{row}+R{row}")
    # O: 원형존치 보전산지 소계 = P + Q (수식)
    ws.cell(row=row, column=15, value=f"=P{row}+Q{row}")
    ws.cell(row=row, column=16, value=0)               # P: 원형존치 임업용
    ws.cell(row=row, column=17, value=0)               # Q: 원형존치 공익용
    ws.cell(row=row, column=18, value=0)               # R: 원형존치 준보전

    # ③ 산지전용협의 = ① 구역지정 - ② 원형존치 (수식)
    ws.cell(row=row, column=19, value=f"=H{row}-I{row}")  # S: ③ 산지전용협의
    # T: 산지전용협의 소유별 합계 = U + V + W (수식)
    ws.cell(row=row, column=20, value=f"=U{row}+V{row}+W{row}")
    ws.cell(row=row, column=21, value=val_guk)         # U: 국유지
    ws.cell(row=row, column=22, value=val_gong)        # V: 공유지
    ws.cell(row=row, column=23, value=val_sa)          # W: 사유지
    # X: 산지전용협의 산지구분별 합계 = Y + AB (수식)
    ws.cell(row=row, column=24, value=f"=Y{row}+AB{row}")
    # Y: 산지전용협의 보전산지 소계 = Z + AA (수식)
    ws.cell(row=row, column=25, value=f"=Z{row}+AA{row}")
    ws.cell(row=row, column=26, value=val_imup)        # Z: 임업용
    ws.cell(row=row, column=27, value=val_gongik)      # AA: 공익용
    ws.cell(row=row, column=28, value=val_junbo)       # AB: 준보전산지

    ws.cell(row=row, column=29, value=owner)           # AC: 소유자
    ws.cell(row=row, column=30, value=inc_area)        # AD: 순증면적
    # AE: 비고 — 비움


def _copy_row_style(ws, src_row, dst_row, max_col):
    """src_row의 셀 서식을 dst_row로 복사 (insert_rows 후 빈 행에 적용)."""
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def create_mountain_report(pnu_results):
    """편입산지조서 생성 — 당초 데이터만, 변경/증감 행 모두 제거.

    Args:
        pnu_results: LandLedgerAnalyzer.analyze() 결과 리스트.

    Returns:
        bytes: 생성된 xlsx 바이너리. 산지 필지가 없으면 None.
    """
    # 템플릿 경로
    template_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "assets", "편입산지조서_template.xlsx",
    )
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"템플릿 파일이 없습니다: {template_path}")

    # 산지 필지 필터링
    mountain_parcels = [
        r for r in (pnu_results or [])
        if r.get("지목") == "임" or r.get("필지구분") == "산"
    ]
    if not mountain_parcels:
        return None

    n_parcels = len(mountain_parcels)

    wb = load_workbook(template_path)
    ws = wb.active
    max_col = ws.max_column  # 33 (A~AG)

    # ⓪ 원본 템플릿의 자동 필터 제거
    #    템플릿에 auto_filter.ref='A9:AG67'이 설정되어 있어, 행 9에 필터 드롭다운이 표시됨.
    #    사용자 요청대로 필터 제거.
    ws.auto_filter.ref = None

    # ① 모든 "변경"/"증감" 행 수집 (B열 기준) — 헤더/합계/데이터 영역 모두 포함
    rows_to_delete = []
    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        if b_val in ("변경", "증감"):
            rows_to_delete.append(row)

    # 아래부터 삭제 (인덱스 변경 회피)
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row, 1)

    # ② 삭제 후 데이터 영역 시작 행 = 원래 행 10 - (행 7,8 삭제) = 행 8
    DATA_START = 8
    template_data_rows = 0
    # 데이터 행 개수 카운트 (B="당초"인 행 또는 비어있는 데이터 행)
    for row in range(DATA_START, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value
        if a_val is not None or b_val == "당초":
            template_data_rows += 1
        else:
            break

    # ③ 필지 수에 맞게 행 추가/삭제
    needed_rows = n_parcels
    if needed_rows > template_data_rows:
        # 행 부족 → 마지막 데이터 행 다음에 추가 (마지막 행 스타일 복사)
        extra = needed_rows - template_data_rows
        insert_at = DATA_START + template_data_rows
        ws.insert_rows(insert_at, extra)
        # 새 행에 직전 행의 스타일 복사
        style_src = DATA_START + template_data_rows - 1
        for i in range(extra):
            _copy_row_style(ws, style_src, insert_at + i, max_col)
    elif needed_rows < template_data_rows:
        # 행 초과 → 사용 안 하는 행 삭제
        remove_count = template_data_rows - needed_rows
        ws.delete_rows(DATA_START + needed_rows, remove_count)

    # ④-a 데이터 영역의 병합 셀 모두 해제
    #     원본 템플릿의 합계 영역(행 7,8)에 있던 E:F 병합과 데이터 행의 변경/증감 행 병합이
    #     delete_rows 후에도 메타데이터로 잔존하여 F열 값 표시를 차단함.
    #     → 데이터 시작 행 이후의 병합을 모두 풀어줌. (헤더 1~6의 병합은 보존)
    merged_to_unmerge = [str(mr) for mr in list(ws.merged_cells.ranges) if mr.min_row >= DATA_START]
    for mr_str in merged_to_unmerge:
        ws.unmerge_cells(mr_str)

    # ④-b 데이터 채우기 — 기존 행의 "당초" 마커도 덮어쓰기
    for idx, parcel in enumerate(mountain_parcels, start=1):
        row = DATA_START + (idx - 1)
        # 기존 셀 값 정리 (B열 "당초" 마커 등)
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            # 수식이 아닌 일반 값만 클리어 (스타일 보존)
            if isinstance(cell.value, str) and cell.value in ("당초", "변경", "증감"):
                cell.value = None
            elif cell.value is not None and not isinstance(cell.value, (int, float, str)):
                cell.value = None
        _fill_data_row(ws, row, idx, parcel)

    # ④-c 데이터 영역 채우기 완전 제거 (어떤 fill 타입이든 무조건 NO_FILL)
    #     원본 템플릿의 "변경" 행에 묻은 파란색이 insert_rows 시 스타일 복사로
    #     사용/미사용 행 모두에 잔존하는 문제를 근원 차단.
    #     ※ 헤더/합계 영역(행 1~6)은 보존.
    used_last_row = DATA_START + n_parcels - 1
    for row in range(DATA_START, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = NO_FILL

    # ④-d 행 단위 fill 제거 (row_dimensions)
    #     원본 템플릿이 "변경" 행에 행 단위(row-level) fill을 설정해 두어,
    #     셀별 NO_FILL을 적용해도 행 색이 위에 표시됨. row_dimensions의 fill도 제거.
    #     ※ 헤더 영역(행 1~6)은 보존.
    for r in list(ws.row_dimensions.keys()):
        if r >= DATA_START:
            rd = ws.row_dimensions[r]
            if hasattr(rd, "fill") and rd.fill is not None:
                rd.fill = NO_FILL

    # ④-e 사용된 마지막 데이터 행 다음의 모든 잔존 행/스타일 완전 삭제.
    #     1) 셀이 있는 빈 행 삭제
    if ws.max_row > used_last_row:
        ws.delete_rows(used_last_row + 1, ws.max_row - used_last_row)
    #     2) row_dimensions 메타데이터도 정리 (delete_rows로 자동 제거되지 않음)
    for r in list(ws.row_dimensions.keys()):
        if r > used_last_row:
            del ws.row_dimensions[r]

    # ④-f 데이터 영역 모든 셀에 일관된 서식 강제 적용
    #     unmerge로 깨진 셀(예: F8 — 원본 E8:F8 병합) 서식 복구.
    #     모든 데이터 셀: 테두리 + 가운데정렬.
    for r in range(DATA_START, used_last_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER_ALIGN

    # ④-g 컬럼 너비 명시 설정 — 템플릿 기본값(C=10.9, E=8)이 좁아 데이터가 잘림
    col_widths = {
        "A": 7,    # 연번
        "B": 7,    # 구분 (빈칸)
        "C": 24,   # 소재지 ★ (예: "경상북도 영천시 신녕면 매양리")
        "D": 6,    # 구분
        "E": 13,   # 지번 ★ (예: "산 1234-5")
        "F": 7,    # 지목 (1글자)
        "G": 12,   # 공부면적
        "H": 12,   # 편입면적
        "AC": 12,  # 소유자
        "AD": 12,  # 순증면적
        "AE": 14,  # 비고
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ⑤ 합계 행 (행 6) — 모든 수치 컬럼에 단순 SUM 수식 작성
    #    (변경/증감 행이 제거되었으므로 원본의 SUMIFS("당초") 필터 → 단순 SUM)
    #    J/N/O/T/X/Y 같은 합계 셀도 데이터 행 수식의 SUM이라 자동으로 정확.
    #    AC(소유자, 텍스트)는 제외.
    last_data_row = DATA_START + n_parcels - 1
    sum_cols = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R",
                "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AD"]
    for col_letter in sum_cols:
        col_idx = column_index_from_string(col_letter)
        ws.cell(
            row=6, column=col_idx,
            value=f"=SUM({col_letter}{DATA_START}:{col_letter}{last_data_row})",
        )
    # AC6은 소유자 텍스트 영역 — 비움
    ws.cell(row=6, column=29, value=None)

    # 합계 행의 E6 셀에 필지 수 표시
    ws.cell(row=6, column=5, value=f"{n_parcels} 필지")
    # B6의 "당초" 마커는 의미 없어졌으므로 "합계"로 변경
    ws.cell(row=6, column=2, value="합계")

    # 시트명 정리 (사업명이 들어가야 할 자리이지만 일단 기본값 유지)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
